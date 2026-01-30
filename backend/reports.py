from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from io import BytesIO
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def generate_booking_pdf(bookings: List[Dict], filters: Dict = {}) -> BytesIO:
    """Generate PDF report for bookings"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=30,
        alignment=1
    )
    title = Paragraph("Booking Report", title_style)
    elements.append(title)
    
    # Date range info
    if filters.get('start_date') or filters.get('end_date'):
        date_info = f"Report Period: {filters.get('start_date', 'N/A')} to {filters.get('end_date', 'N/A')}"
        elements.append(Paragraph(date_info, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    data = [['PNR', 'Passenger', 'Airline', 'Sale Price', 'Status']]
    for booking in bookings:
        data.append([
            booking.get('pnr', 'N/A'),
            booking.get('pax_name', 'N/A'),
            booking.get('airline', 'N/A'),
            f"${booking.get('sale_price', 0):.2f}",
            booking.get('status', 'N/A').replace('_', ' ').title()
        ])
    
    table = Table(data, colWidths=[1.2*inch, 1.8*inch, 1.2*inch, 1*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_booking_excel(bookings: List[Dict], filters: Dict = {}) -> BytesIO:
    """Generate Excel report for bookings"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"
    
    # Header style
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    # Headers
    headers = ['PNR', 'Passenger Name', 'Contact', 'Airline', 'Supplier', 'Our Cost', 'Sale Price', 'Status', 'Created By', 'Created At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row, booking in enumerate(bookings, 2):
        ws.cell(row=row, column=1, value=booking.get('pnr', ''))
        ws.cell(row=row, column=2, value=booking.get('pax_name', ''))
        ws.cell(row=row, column=3, value=booking.get('contact_number', ''))
        ws.cell(row=row, column=4, value=booking.get('airline', ''))
        ws.cell(row=row, column=5, value=booking.get('supplier_name', 'N/A'))
        ws.cell(row=row, column=6, value=booking.get('our_cost', 0))
        ws.cell(row=row, column=7, value=booking.get('sale_price', 0))
        ws.cell(row=row, column=8, value=booking.get('status', '').replace('_', ' ').title())
        ws.cell(row=row, column=9, value=booking.get('created_by_name', ''))
        ws.cell(row=row, column=10, value=booking.get('created_at', ''))
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_outstanding_balance_excel(report_data: List[Dict]) -> BytesIO:
    """Generate Excel report for outstanding balances"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Outstanding Balances"
    
    # Header style
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    # Headers
    headers = ['PNR', 'Passenger', 'Supplier', 'Sale Price', 'Total Paid', 'Balance', 'Created At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    total_balance = 0
    for row, item in enumerate(report_data, 2):
        ws.cell(row=row, column=1, value=item.get('pnr', ''))
        ws.cell(row=row, column=2, value=item.get('pax_name', ''))
        ws.cell(row=row, column=3, value=item.get('supplier_name', ''))
        ws.cell(row=row, column=4, value=item.get('sale_price', 0))
        ws.cell(row=row, column=5, value=item.get('total_paid', 0))
        balance = item.get('balance', 0)
        ws.cell(row=row, column=6, value=balance)
        ws.cell(row=row, column=7, value=item.get('created_at', ''))
        total_balance += balance
    
    # Total row
    total_row = len(report_data) + 2
    ws.cell(row=total_row, column=5, value="Total Outstanding:")
    ws.cell(row=total_row, column=6, value=total_balance)
    ws.cell(row=total_row, column=6).font = Font(bold=True, color="EF4444")
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
